#!/usr/bin/env python3
"""
localize_special_topic_page.py
将 special-topic-page 从英文版创建指定语言的 localization。

用法:
    python3 localize_special_topic_page.py <en_page_id> <locale> <excel_sheet_name>

示例:
    python3 localize_special_topic_page.py 5 fr "Organize PDF"

逻辑:
  1. 拉取英文版完整结构（含所有嵌套字段、图片 id、buttons 等）
  2. 读取 Excel 对应 sheet，提取 English → 目标语言 的翻译映射
  3. 对英文版 blocks 进行文本替换，图片/按钮/steps/faq id 等全部复用
  4. POST /api/special-topic-pages/{id}/localizations 创建新 locale
  5. PUT publishedAt 发布
"""

import os
import sys
import json
import copy
import argparse
import requests
import openpyxl

# ── 配置 ─────────────────────────────────────────────────────────────────────

CMS_TOKEN = os.environ.get("CMS_TOKEN_TEST", "")
CMS_BASE = os.environ.get("CMS_BASE_TEST", "http://pdfagile-cms.aix-test-k8s.iweikan.cn")

EXCEL_PATH = os.environ.get("LOCALIZE_EXCEL", "")

# locale code → Excel 列标题映射
LOCALE_TO_EXCEL_COL = {
    "fr":      "French",
    "zh-Hant": "Traditional Chinese",
    "es":      "Spanish",
    "de":      "German",
    "pt":      "Portuguese",
    "it":      "Italian",
    "ja":      "Japanese",
    "ko":      "Korean",
    "ar":      "Arabic",
    "id":      "Indonesian",
    "vi":      "Vietnamese",
    "th":      "Thai",
    "ms":      "Malay",
    "tr":      "Turkish",
    "pl":      "Polish",
    "nl":      "Dutch",
    "ro":      "Romanian",
    "hi":      "Hindi",
}

# ── CMS helpers ───────────────────────────────────────────────────────────────

def headers():
    return {"Authorization": f"Bearer {CMS_TOKEN}"}

def headers_json():
    return {**headers(), "Content-Type": "application/json"}


def fetch_en_page(page_id: int) -> dict:
    """拉取英文版完整结构，包含 blocks 所有子字段（image、buttons、steps、faq、trustedBy icon 等）"""
    url = (
        f"{CMS_BASE}/api/special-topic-pages/{page_id}"
        "?populate[blocks][populate][trustedBy][populate]=*"
        "&populate[blocks][populate][media]=*"
        "&populate[blocks][populate][backgroundImage]=*"
        "&populate[blocks][populate][icon]=*"
        "&populate[blocks][populate][buttons][populate]=*"
        "&populate[blocks][populate][background]=*"
        "&populate[blocks][populate][header][populate][icon]=*"
        "&populate[blocks][populate][steps][populate]=*"
        "&populate[blocks][populate][faq]=*"
        "&populate[blocks][populate][subFeatures][populate]=*"
        "&populate[seo][populate]=*"
        "&locale=en"
    )
    resp = requests.get(url, headers=headers())
    resp.raise_for_status()
    data = resp.json()
    return data["data"]["attributes"]


# ── Excel translation map ─────────────────────────────────────────────────────

# 全局 FAQ 翻译结构：{ en_question_stripped: fr_answer_html }
# 由 build_translation_map 填充，供 strip_faq 使用
_FAQ_ANSWER_MAP: dict = {}


def build_translation_map(sheet_name: str, locale: str) -> dict:
    """
    读取 Excel sheet，返回 {english_text: translated_text} 映射。
    同时填充全局 _FAQ_ANSWER_MAP 用于 FAQ answer 翻译。
    """
    col_name = LOCALE_TO_EXCEL_COL.get(locale)
    if not col_name:
        raise ValueError(f"不支持的 locale: {locale}，支持的有: {list(LOCALE_TO_EXCEL_COL.keys())}")

    if not os.path.exists(EXCEL_PATH):
        raise FileNotFoundError(f"找不到 Excel 文件: {EXCEL_PATH}")

    wb = openpyxl.load_workbook(EXCEL_PATH)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Excel 中找不到 sheet: '{sheet_name}'，可用的有: {wb.sheetnames}")

    ws = wb[sheet_name]
    headers_row = [cell.value for cell in ws[1]]

    # 找列索引
    try:
        en_col = headers_row.index("English")
    except ValueError:
        raise ValueError("Excel sheet 中找不到 'English' 列")
    try:
        tr_col = headers_row.index(col_name)
    except ValueError:
        raise ValueError(f"Excel sheet 中找不到 '{col_name}' 列")

    # Notes 列（用于 FAQ 分组）
    notes_col = None
    if "Notes" in headers_row:
        notes_col = headers_row.index("Notes")

    all_rows = list(ws.iter_rows(min_row=2, values_only=True))

    translation_map = {}
    # 普通一对一映射
    for row in all_rows:
        en = row[en_col] if len(row) > en_col else None
        tr = row[tr_col] if len(row) > tr_col else None
        if en and tr and str(en).strip() and str(tr).strip():
            translation_map[str(en).strip()] = str(tr).strip()

    # FAQ 分组：按 Q 行切分，拼接 answer HTML，写入全局 _FAQ_ANSWER_MAP
    global _FAQ_ANSWER_MAP
    _FAQ_ANSWER_MAP = {}
    if notes_col is not None:
        import re as _re
        faq_rows_data = [
            (row[en_col], row[tr_col])
            for row in all_rows
            if row[notes_col] and "FAQ" in str(row[notes_col])
               and row[en_col] and row[tr_col]
        ]
        # 切分为 [(en_q, fr_q, [(en_a_line, fr_a_line), ...]), ...]
        groups: list = []
        cur_q = None
        for en_text, fr_text in faq_rows_data:
            en_s = str(en_text).strip()
            fr_s = str(fr_text).strip()
            if _re.match(r'^Q\d+[:\s]', en_s):
                en_q_stripped = _re.sub(r'^Q\d+[:\s]+', '', en_s).strip()
                fr_q_stripped = _re.sub(r'^Q\d+[\s\uff1a:]+', '', fr_s).strip()
                cur_q = (en_q_stripped, fr_q_stripped, [])
                groups.append(cur_q)
                # 也加入普通翻译表（question 文本）
                translation_map[en_q_stripped] = fr_q_stripped
            elif cur_q is not None:
                cur_q[2].append((en_s, fr_s))

        for (en_q_stripped, fr_q_stripped, answer_lines) in groups:
            if answer_lines:
                fr_answer_html = "".join(f"<p>{line[1]}</p>" for line in answer_lines)
                _FAQ_ANSWER_MAP[en_q_stripped] = fr_answer_html

    # 补丁：Strapi 里的文本与 Excel 有细微差异的条目，按 locale 手动对齐
    # 格式: { locale: { en_text: translated_text } }
    MANUAL_PATCHES: dict = {
        "fr": {
            "How to Rearrange a PDF document in 3 easy steps?":
                "Comment réorganiser des pages PDF en 3 étapes faciles ?",
            "Accelerate your PDF workflow like never before":
                "Accélérez votre flux de travail PDF comme jamais auparavant",
            "A complete PDF solution for all PDF needs":
                "Une solution PDF complète pour tous les besoins en matière de PDF",
            "Free Download": "Téléchargement gratuit",
            "Pricing":       "Tarifs",
        },
        "zh-Hant": {
            "How to Rearrange a PDF document in 3 easy steps?":
                "如何用3個簡單步驟重新排列PDF文件？",
            "Accelerate your PDF workflow like never before":
                "以前所未有的方式加速您的PDF工作流程",
            "A complete PDF solution for all PDF needs":
                "滿足所有PDF需求的完整解決方案",
            "Free Download": "免費下載",
            "Pricing":       "價格方案",
        },
        "es": {
            "How to Rearrange a PDF document in 3 easy steps?":
                "¿Cómo reorganizar un documento PDF en 3 sencillos pasos?",
            "Accelerate your PDF workflow like never before":
                "Acelera tu flujo de trabajo PDF como nunca antes",
            "A complete PDF solution for all PDF needs":
                "Una solución PDF completa para todas tus necesidades",
            "Free Download": "Descarga gratis",
            "Pricing":       "Precios",
        },
        "de": {
            "How to Rearrange a PDF document in 3 easy steps?":
                "Wie ordnet man PDF-Seiten in 3 einfachen Schritten neu an?",
            "Accelerate your PDF workflow like never before":
                "Beschleunigen Sie Ihren PDF-Workflow wie nie zuvor",
            "A complete PDF solution for all PDF needs":
                "Eine vollständige PDF-Lösung für alle PDF-Anforderungen",
            "Free Download": "Kostenlos herunterladen",
            "Pricing":       "Preise",
        },
        "pt": {
            "How to Rearrange a PDF document in 3 easy steps?":
                "Como reorganizar páginas PDF em 3 etapas fáceis?",
            "Accelerate your PDF workflow like never before":
                "Acelere seu fluxo de trabalho PDF como nunca antes",
            "A complete PDF solution for all PDF needs":
                "Uma solução PDF completa para todas as necessidades",
            "Free Download": "Download grátis",
            "Pricing":       "Preços",
        },
        "it": {
            "How to Rearrange a PDF document in 3 easy steps?":
                "Come riorganizzare le pagine PDF in 3 semplici passaggi?",
            "Accelerate your PDF workflow like never before":
                "Accelera il tuo flusso di lavoro PDF come mai prima",
            "A complete PDF solution for all PDF needs":
                "Una soluzione PDF completa per tutte le esigenze",
            "Free Download": "Download gratuito",
            "Pricing":       "Prezzi",
        },
        "ja": {
            "How to Rearrange a PDF document in 3 easy steps?":
                "3つの簡単なステップでPDFページを並べ替える方法",
            "Accelerate your PDF workflow like never before":
                "これまでにないほどPDFワークフローを加速させましょう",
            "A complete PDF solution for all PDF needs":
                "あらゆるPDFニーズに対応する完全なPDFソリューション",
            "Free Download": "無料ダウンロード",
            "Pricing":       "料金",
        },
        "ko": {
            "How to Rearrange a PDF document in 3 easy steps?":
                "3가지 간단한 단계로 PDF 페이지를 재배열하는 방법",
            "Accelerate your PDF workflow like never before":
                "전례 없는 속도로 PDF 워크플로우를 가속하세요",
            "A complete PDF solution for all PDF needs":
                "모든 PDF 요구 사항을 위한 완벽한 PDF 솔루션",
            "Free Download": "무료 다운로드",
            "Pricing":       "가격",
        },
        "ar": {
            "How to Rearrange a PDF document in 3 easy steps?":
                "كيف تعيد ترتيب صفحات PDF في 3 خطوات سهلة؟",
            "Accelerate your PDF workflow like never before":
                "سرّع سير عمل PDF الخاص بك كما لم يحدث من قبل",
            "A complete PDF solution for all PDF needs":
                "حل PDF متكامل لجميع احتياجات PDF",
            "Free Download": "تحميل مجاني",
            "Pricing":       "الأسعار",
        },
        "id": {
            "How to Rearrange a PDF document in 3 easy steps?":
                "Bagaimana cara menyusun ulang halaman PDF dalam 3 langkah mudah?",
            "Accelerate your PDF workflow like never before":
                "Percepat alur kerja PDF Anda seperti belum pernah sebelumnya",
            "A complete PDF solution for all PDF needs":
                "Solusi PDF lengkap untuk semua kebutuhan PDF",
            "Free Download": "Unduh gratis",
            "Pricing":       "Harga",
        },
        "vi": {
            "How to Rearrange a PDF document in 3 easy steps?":
                "Làm thế nào để sắp xếp lại các trang PDF trong 3 bước đơn giản?",
            "Accelerate your PDF workflow like never before":
                "Tăng tốc quy trình làm việc PDF của bạn như chưa từng có",
            "A complete PDF solution for all PDF needs":
                "Giải pháp PDF hoàn chỉnh cho mọi nhu cầu PDF",
            "Free Download": "Tải xuống miễn phí",
            "Pricing":       "Bảng giá",
        },
        "th": {
            "How to Rearrange a PDF document in 3 easy steps?":
                "วิธีการจัดเรียงหน้า PDF ใหม่ใน 3 ขั้นตอนง่ายๆ",
            "Accelerate your PDF workflow like never before":
                "เร่งความเร็วเวิร์กโฟลว์ PDF ของคุณอย่างที่ไม่เคยมีมาก่อน",
            "A complete PDF solution for all PDF needs":
                "โซลูชัน PDF ที่ครบครันสำหรับทุกความต้องการด้าน PDF",
            "Free Download": "ดาวน์โหลดฟรี",
            "Pricing":       "ราคา",
        },
        "ms": {
            "How to Rearrange a PDF document in 3 easy steps?":
                "Bagaimana untuk menyusun semula halaman PDF dalam 3 langkah mudah?",
            "Accelerate your PDF workflow like never before":
                "Percepatkan aliran kerja PDF anda seperti tidak pernah sebelumnya",
            "A complete PDF solution for all PDF needs":
                "Penyelesaian PDF lengkap untuk semua keperluan PDF",
            "Free Download": "Muat turun percuma",
            "Pricing":       "Harga",
        },
        "tr": {
            "How to Rearrange a PDF document in 3 easy steps?":
                "PDF sayfaları 3 kolay adımda nasıl yeniden düzenlenir?",
            "Accelerate your PDF workflow like never before":
                "PDF iş akışınızı hiç olmadığı kadar hızlandırın",
            "A complete PDF solution for all PDF needs":
                "Tüm PDF ihtiyaçları için eksiksiz bir PDF çözümü",
            "Free Download": "Ücretsiz indir",
            "Pricing":       "Fiyatlandırma",
        },
        "pl": {
            "How to Rearrange a PDF document in 3 easy steps?":
                "Jak zmienić kolejność stron PDF w 3 prostych krokach?",
            "Accelerate your PDF workflow like never before":
                "Przyspiesz swój przepływ pracy z PDF jak nigdy dotąd",
            "A complete PDF solution for all PDF needs":
                "Kompleksowe rozwiązanie PDF dla wszystkich potrzeb PDF",
            "Free Download": "Pobierz za darmo",
            "Pricing":       "Cennik",
        },
        "nl": {
            "How to Rearrange a PDF document in 3 easy steps?":
                "Hoe herschik ik PDF-pagina's in 3 eenvoudige stappen?",
            "Accelerate your PDF workflow like never before":
                "Versnel uw PDF-workflow zoals nooit tevoren",
            "A complete PDF solution for all PDF needs":
                "Een complete PDF-oplossing voor alle PDF-behoeften",
            "Free Download": "Gratis downloaden",
            "Pricing":       "Prijzen",
        },
        "ro": {
            "How to Rearrange a PDF document in 3 easy steps?":
                "Cum să reorganizezi paginile PDF în 3 pași simpli?",
            "Accelerate your PDF workflow like never before":
                "Accelerează fluxul tău de lucru PDF ca niciodată",
            "A complete PDF solution for all PDF needs":
                "O soluție PDF completă pentru toate nevoile PDF",
            "Free Download": "Descărcare gratuită",
            "Pricing":       "Prețuri",
        },
        "hi": {
            "How to Rearrange a PDF document in 3 easy steps?":
                "3 आसान चरणों में PDF पृष्ठों को कैसे पुनर्व्यवस्थित करें?",
            "Accelerate your PDF workflow like never before":
                "अपने PDF वर्कफ़्लो को पहले से कहीं अधिक तेज़ करें",
            "A complete PDF solution for all PDF needs":
                "सभी PDF आवश्यकताओं के लिए एक संपूर्ण PDF समाधान",
            "Free Download": "मुफ्त डाउनलोड",
            "Pricing":       "मूल्य निर्धारण",
        },
    }
    for en_key, tr_val in MANUAL_PATCHES.get(locale, {}).items():
        if en_key not in translation_map:
            translation_map[en_key] = tr_val

    print(f"  翻译表加载完成：{len(translation_map)} 条普通映射，{len(_FAQ_ANSWER_MAP)} 条 FAQ 分组（sheet: {sheet_name}）")
    return translation_map


def translate(text, t_map: dict) -> str:
    """
    精确匹配翻译，找不到则返回原文。
    标准化处理：去首尾空白、合并连续空格。
    同时尝试去掉尾部句号后匹配（应对源文本句号不一致问题）。
    """
    if not text:
        return text
    import re
    key = re.sub(r'\s+', ' ', str(text).strip())
    for k, v in t_map.items():
        k_norm = re.sub(r'\s+', ' ', str(k).strip())
        if k_norm == key:
            return v
    # 尝试去掉尾部句号再匹配
    key_stripped = key.rstrip('.')
    if key_stripped != key:
        for k, v in t_map.items():
            k_norm = re.sub(r'\s+', ' ', str(k).strip())
            if k_norm == key_stripped:
                return v
    return text


# ── Block 转换：文本替换，图片/按钮/关联 id 复用 ──────────────────────────────

def strip_media_to_id(field_val):
    """
    将 {"data": {"id": 123, "attributes": {...}}} 转为 123（Strapi 写入时只需要 id）。
    {"data": null} → None
    """
    if isinstance(field_val, dict):
        data = field_val.get("data")
        if data is None:
            return None
        if isinstance(data, dict):
            return data.get("id")
    return field_val


class VarcharTooLongError(Exception):
    """字段超过 Strapi VARCHAR(255) 限制时抛出，携带字段名和实际长度供上层展示。"""
    def __init__(self, field: str, length: int, limit: int = 255):
        self.field   = field
        self.length  = length
        self.limit   = limit
        super().__init__(
            f"字段 `{field}` 长度 {length} 超过 Strapi VARCHAR({limit}) 限制。"
            f"建议在 Strapi Content-Type Builder 中将该字段类型改为 Long text（Text）。"
        )


def check_varchar(text: str, field: str, limit: int = 255, force_truncate: bool = False) -> str:
    """检查字符串是否超过 VARCHAR 限制。
    force_truncate=True 时截断并继续；否则抛出 VarcharTooLongError。
    """
    if text is None or len(text) <= limit:
        return text
    if force_truncate:
        cut = text.rfind(" ", 0, limit)
        if cut < limit // 2:
            cut = limit
        return text[:cut]
    raise VarcharTooLongError(field, len(text), limit)


def truncate_varchar(text: str, limit: int = 255) -> str:
    """兼容旧调用：直接截断（仅在 force_truncate 场景使用）。"""
    return check_varchar(text, "unknown", limit, force_truncate=True)


def strip_buttons(buttons: list) -> list:
    """
    将 buttons 列表转为可写入格式，保留结构，去掉只读属性。
    支持两种格式：
      - feature-hero button: {id, theme, href, label, target, isExternal, disabled}
      - cta button: {id, theme, link: {href, label, ...}}
    """
    if not buttons:
        return []
    result = []
    for btn in buttons:
        if "link" in btn:
            # cta 格式：只传 id（复用已有记录）
            result.append({"id": btn["id"], "theme": btn["theme"]})
        else:
            # feature-hero 格式
            b = {
                "theme":      btn.get("theme"),
                "href":       btn.get("href"),
                "label":      btn.get("label"),
                "target":     btn.get("target"),
                "isExternal": btn.get("isExternal", False),
                "disabled":   btn.get("disabled", False),
            }
            if btn.get("image"):
                img_id = strip_media_to_id(btn["image"])
                if img_id:
                    b["image"] = img_id
            result.append(b)
    return result


def strip_steps(steps: list, t_map: dict) -> list:
    """step-cards 的 steps 子项：翻译 customizeText（去 HTML 标签后匹配），复用 media id"""
    if not steps:
        return []
    import re as _re
    result = []
    for s in steps:
        # customizeText 格式: "<p>Launch PDF Agile...</p>" (Strapi 无编号)
        # Excel 里对应行格式: "1. Launch PDF Agile..." (有编号前缀)
        # 策略：先直接匹配，再尝试在 t_map 里找带编号前缀的 key
        raw_ct = s.get("customizeText") or ""
        plain_ct = _re.sub(r'<[^>]+>', '', raw_ct).strip()
        fr_ct = translate(plain_ct, t_map)
        if fr_ct == plain_ct:
            # 尝试在翻译表里找 "N. {plain_ct}" 格式的 key
            for k, v in t_map.items():
                k_body = _re.sub(r'^\d+\.\s*', '', k.strip())
                if k_body == plain_ct:
                    # 取对应法语，去掉编号前缀
                    fr_ct = _re.sub(r'^\d+[\.\s：:]+\s*', '', v.strip())
                    break
        if fr_ct and fr_ct != plain_ct:
            translated_ct = f"<p>{fr_ct}</p>" if not fr_ct.startswith("<") else fr_ct
        else:
            translated_ct = raw_ct  # 找不到翻译保留原文
        step = {
            "layout":        s.get("layout"),
            "theme":         s.get("theme"),
            "text":          translate(s.get("text"), t_map),
            "customizeText": translated_ct,
            "title":         translate(s.get("title"), t_map),
            "customizeTitle": s.get("customizeTitle", ""),
            "topTitle":      s.get("topTitle"),
        }
        # 复用 media
        if s.get("media"):
            mid = strip_media_to_id(s["media"])
            if mid:
                step["media"] = mid
        result.append(step)
    return result


def strip_faq(faq_items: list, t_map: dict) -> list:
    """
    faq 条目：翻译 question 和 answer。
    - question：精确文本匹配翻译
    - answer：用 _FAQ_ANSWER_MAP（按 question key 查对应 FR answer HTML）
              找不到时保留英文原文
    """
    if not faq_items:
        return []
    result = []
    for item in faq_items:
        import re
        raw_q = item.get("question", "")
        q = translate(raw_q, t_map)

        # answer：通过 question 找对应的 FR answer HTML
        raw_ans = item.get("answer", "")
        q_key = re.sub(r'\s+', ' ', str(raw_q).strip())
        fr_ans = _FAQ_ANSWER_MAP.get(q_key)
        if fr_ans:
            translated_ans = fr_ans
        else:
            # 回退：尝试普通翻译表（通常找不到，保留英文）
            translated_ans = translate(raw_ans, t_map)
        result.append({"question": q, "answer": translated_ans})
    return result


def strip_trusted_by(items: list) -> list:
    """trust-by 的 trustedBy 子项：直接复用（品牌 label 不翻译）"""
    if not items:
        return []
    return [{"id": item["id"], "label": item["label"]} for item in items]


def convert_block(block: dict, t_map: dict, locale: str, force_truncate: bool = False) -> dict:
    """
    将英文版 block 转换为目标语言版本：
    - 文本字段：查翻译表，找不到保留英文
    - media/image 字段：提取 id 复用
    - buttons/steps/faq 等嵌套结构：专项处理
    - 只读字段（id）：去掉（由 Strapi 自动生成）
    """
    comp = block["__component"]
    b = {"__component": comp}

    if comp == "feature.feature-hero":
        # 只传非 null 的可选字段，避免 Strapi 500
        top = translate(block.get("topTitle"), t_map)
        if top is not None:
            b["topTitle"] = top
        b["title"]    = translate(block.get("title"), t_map)
        b["subtitle"] = check_varchar(translate(block.get("subtitle"), t_map), "feature-hero.subtitle", force_truncate=force_truncate)
        if block.get("layout") is not None:
            b["layout"] = block["layout"]
        if block.get("jsonData") is not None:
            b["jsonData"] = block["jsonData"]
        if block.get("theme") is not None:
            b["theme"] = block["theme"]
        # 复用 backgroundImage
        bg_id = strip_media_to_id(block.get("backgroundImage"))
        if bg_id:
            b["backgroundImage"] = {"id": bg_id}
        # 翻译 buttons 的 label，其余字段保留（不传 image 字段，避免 Strapi 校验报错）
        raw_btns = block.get("buttons", [])
        new_btns = []
        for btn in raw_btns:
            new_btn = {
                "theme":      btn.get("theme"),
                "href":       btn.get("href"),
                "label":      translate(btn.get("label"), t_map),
                "target":     btn.get("target"),
                "isExternal": btn.get("isExternal", False),
                "disabled":   btn.get("disabled", False),
            }
            new_btns.append(new_btn)
        b["buttons"] = new_btns

    elif comp == "feature.panel":
        b["layout"]          = block.get("layout")
        b["theme"]           = block.get("theme")
        b["text"]            = block.get("text")
        b["customizeText"]   = translate(
            # customizeText 带 HTML 标签，去掉后匹配
            block.get("customizeText", "").replace("<p>", "").replace("</p>", "").strip()
            if block.get("customizeText") else block.get("customizeText"),
            t_map
        )
        # 如果翻译表找到的是纯文本，包回 <p> 标签
        if b["customizeText"] and not b["customizeText"].startswith("<"):
            b["customizeText"] = f"<p>{b['customizeText']}</p>"
        b["title"]           = translate(block.get("title"), t_map)
        b["customizeTitle"]  = block.get("customizeTitle", "")
        b["topTitle"]        = translate(block.get("topTitle"), t_map)
        # 复用 media / backgroundImage / icon
        media_id = strip_media_to_id(block.get("media"))
        if media_id:
            b["media"] = {"id": media_id}
        bg_id = strip_media_to_id(block.get("backgroundImage"))
        if bg_id:
            b["backgroundImage"] = {"id": bg_id}
        icon_id = strip_media_to_id(block.get("icon"))
        if icon_id:
            b["icon"] = icon_id

    elif comp == "feature.specific-features":
        # subFeatures 是内嵌 component，需要翻译 title 和 text，不传 id（新建）
        subs = []
        for sf in block.get("subFeatures", []):
            sub = {
                "layout":         sf.get("layout"),
                "text":           check_varchar(translate(sf.get("text"), t_map), "subFeatures.text", force_truncate=force_truncate),
                "customizeText":  sf.get("customizeText", ""),
                "title":          translate(sf.get("title"), t_map),
                "customizeTitle": sf.get("customizeTitle", ""),
            }
            # null 可选字段不传，避免 Strapi VARCHAR 校验报 500
            if sf.get("theme") is not None:
                sub["theme"] = sf["theme"]
            if sf.get("topTitle") is not None:
                sub["topTitle"] = sf["topTitle"]
            media_id = strip_media_to_id(sf.get("media"))
            if media_id:
                sub["media"] = {"id": media_id}
            subs.append(sub)
        b["subFeatures"] = subs
        if block.get("mainFeature"):
            b["mainFeature"] = block["mainFeature"]
        bg_id = strip_media_to_id(block.get("backgroundImage"))
        if bg_id:
            b["backgroundImage"] = {"id": bg_id}

    elif comp == "feature.step-cards":
        b["title"]    = translate(block.get("title"), t_map)
        b["theme"]    = block.get("theme")
        b["subtitle"] = check_varchar(translate(block.get("subtitle"), t_map), "step-cards.subtitle", force_truncate=force_truncate)
        b["steps"]    = strip_steps(block.get("steps", []), t_map)

    elif comp == "feature.trust-by":
        b["title"]  = translate(block.get("title"), t_map)
        b["label"]  = block.get("label")
        bg_id = strip_media_to_id(block.get("backgroundImage"))
        if bg_id:
            b["backgroundImage"] = {"id": bg_id}
        # trustedBy 内嵌 component 在 POST localization 时不能传 id（报500），不传即可
        # 后续通过 patch_localization 用 PUT 补全（PUT 同样不支持），故留空

    elif comp == "blocks.faq":
        b["title"]          = translate(block.get("title"), t_map)
        b["theme"]          = block.get("theme")
        b["customizeTitle"] = block.get("customizeTitle", "")
        b["faq"]            = strip_faq(block.get("faq", []), t_map)
        bg_id = strip_media_to_id(block.get("backgroundImage"))
        if bg_id:
            b["backgroundImage"] = {"id": bg_id}

    elif comp == "blocks.cta":
        b["theme"] = block.get("theme")
        b["text"]  = block.get("text")
        bg_id = strip_media_to_id(block.get("background"))
        if bg_id:
            b["background"] = {"id": bg_id}
        # header：新建（不传 id，POST/PUT localization 不允许跨实体复用 component id）
        if block.get("header"):
            h = block["header"]
            b["header"] = {
                "theme":          h.get("theme"),
                "label":          h.get("label"),
                "title":          h.get("title"),
                "customizeTitle": h.get("customizeTitle", ""),
                "customizeText":  h.get("customizeText", ""),
            }
        # buttons：新建（不传 id）
        b["buttons"] = [{"theme": btn["theme"]} for btn in block.get("buttons", [])]

    elif comp == "feature.swiper":
        b["title"] = translate(block.get("title"), t_map)
        b["theme"] = block.get("theme")
        bg_id = strip_media_to_id(block.get("backgroundImage"))
        if bg_id:
            b["backgroundImage"] = {"id": bg_id}
        # swiper items（若有）
        items = block.get("items") or block.get("swiperItems") or []
        if items:
            new_items = []
            for item in items:
                ni = {k: v for k, v in item.items() if k not in ("id", "backgroundImage", "media", "icon")}
                ni["title"] = translate(item.get("title"), t_map)
                ni["text"]  = translate(item.get("text"), t_map)
                media_id = strip_media_to_id(item.get("media"))
                if media_id:
                    ni["media"] = {"id": media_id}
                new_items.append(ni)
            b["items"] = new_items

    else:
        # 未知 component：原样复制，去掉 id，并清理 media 字段
        MEDIA_FIELDS = ("backgroundImage", "media", "icon", "background", "image")
        for k, v in block.items():
            if k == "id":
                continue
            if k in MEDIA_FIELDS:
                cleaned = strip_media_to_id(v)
                if cleaned is not None:
                    b[k] = cleaned
                # None 不传（避免 Strapi 报500）
            else:
                b[k] = v

    return b


# ── 主流程 ────────────────────────────────────────────────────────────────────

def fetch_fr_blocks(new_id: int) -> list:
    """拉取已创建的 localization 的 blocks（带 id），用于后续 PUT 补全"""
    url = (
        f"{CMS_BASE}/api/special-topic-pages/{new_id}"
        "?populate[blocks][populate]=*"
    )
    resp = requests.get(url, headers=headers())
    resp.raise_for_status()
    return resp.json()["data"]["attributes"].get("blocks", [])


def build_patch_blocks(fr_blocks: list, en_blocks: list, t_map: dict, force_truncate: bool = False) -> list:
    """
    构建 PUT 用的完整 blocks 数组：
    - 保留所有已创建 block 的 id
    - 对 feature.trust-by 补入 trustedBy（从英文版取 label 列表，新建不传 id）
    - 其他 block 只传必要字段 + id，保持不变
    """
    # 英文版 trust-by 的 trustedBy：带 label + icon media id（复用图片）
    en_trusted_labels = []
    for b in en_blocks:
        if b["__component"] == "feature.trust-by":
            for t in b.get("trustedBy", []):
                item = {"label": t["label"]}
                icon_id = strip_media_to_id(t.get("icon"))
                if icon_id:
                    item["icon"] = icon_id
                en_trusted_labels.append(item)
            break

    result = []
    for fb in fr_blocks:
        comp = fb["__component"]
        bid = fb["id"]

        if comp == "feature.trust-by":
            entry = {
                "id":        bid,
                "__component": comp,
                "title":     fb.get("title"),
                "label":     fb.get("label"),
                "trustedBy": en_trusted_labels,  # 补入品牌列表（含 icon）
            }
            bg_id = strip_media_to_id(fb.get("backgroundImage"))
            if bg_id:
                entry["backgroundImage"] = {"id": bg_id}
            result.append(entry)

        elif comp == "feature.specific-features":
            # subFeatures 从英文版重建（法语版创建时为空），翻译 title/text，复用 media id
            en_sf_block = next((b for b in en_blocks if b["__component"] == "feature.specific-features"), None)
            en_subs = en_sf_block.get("subFeatures", []) if en_sf_block else []
            subs = []
            for sf in en_subs:
                sub = {
                    "layout":        sf.get("layout"),
                    "text":          check_varchar(translate(sf.get("text"), t_map), "subFeatures.text", force_truncate=force_truncate),
                    "customizeText": sf.get("customizeText", ""),
                    "title":         translate(sf.get("title"), t_map),
                    "customizeTitle": sf.get("customizeTitle", ""),
                }
                if sf.get("theme") is not None:
                    sub["theme"] = sf["theme"]
                if sf.get("topTitle") is not None:
                    sub["topTitle"] = sf["topTitle"]
                media_id = strip_media_to_id(sf.get("media"))
                if media_id:
                    sub["media"] = {"id": media_id}
                subs.append(sub)
            result.append({
                "id":           bid,
                "__component":  comp,
                "subFeatures":  subs,
            })

        elif comp == "blocks.cta":
            # 从英文版取 cta 结构（buttons 含 link）
            en_cta = next((b for b in en_blocks if b["__component"] == "blocks.cta"), None)
            entry = {
                "id":       bid,
                "__component": comp,
                "theme":    fb.get("theme"),
                "text":     fb.get("text"),
            }
            bg_id = strip_media_to_id(fb.get("background"))
            if bg_id:
                entry["background"] = {"id": bg_id}
            if fb.get("header"):
                h = fb["header"]
                # 从英文版 header 取原文翻译，icon 复用
                en_cta_header = (en_cta or {}).get("header") or {}
                icon_id = strip_media_to_id(en_cta_header.get("icon"))
                header_entry = {
                    "id":             h["id"],
                    "theme":          en_cta_header.get("theme") or h.get("theme"),
                    "label":          translate(en_cta_header.get("label") or h.get("label"), t_map),
                    "title":          translate(en_cta_header.get("title") or h.get("title"), t_map),
                    "customizeTitle": h.get("customizeTitle", ""),
                    "customizeText":  h.get("customizeText", ""),
                }
                if icon_id:
                    header_entry["icon"] = icon_id
                entry["header"] = header_entry
            # buttons：带 link，从英文版取 href/target/isExternal，翻译 label
            fr_btns = fb.get("buttons", [])
            en_btns = en_cta.get("buttons", []) if en_cta else []
            patched_buttons = []
            for i, btn in enumerate(fr_btns):
                en_btn = en_btns[i] if i < len(en_btns) else {}
                en_link = en_btn.get("link") or {}
                btn_entry = {"id": btn["id"], "theme": btn["theme"]}
                if en_link:
                    btn_entry["link"] = {
                        "href":       en_link.get("href"),
                        "label":      translate(en_link.get("label"), t_map),
                        "target":     en_link.get("target"),
                        "isExternal": en_link.get("isExternal", False),
                        "disabled":   en_link.get("disabled", False),
                    }
                patched_buttons.append(btn_entry)
            entry["buttons"] = patched_buttons
            result.append(entry)

        elif comp == "blocks.faq":
            # 重新翻译 FAQ（POST 时可能用了旧逻辑，此处用最新的 _FAQ_ANSWER_MAP）
            # 需要从英文版拿原始 question/answer（fr_blocks 里 question 已是法语，无法反向查）
            en_faq_block = next((b for b in en_blocks if b["__component"] == "blocks.faq"), None)
            if en_faq_block:
                faq_translated = strip_faq(en_faq_block.get("faq", []), t_map)
            else:
                faq_translated = fb.get("faq", [])
            entry = {
                "id":           bid,
                "__component":  comp,
                "title":        translate(en_faq_block.get("title") if en_faq_block else fb.get("title"), t_map),
                "theme":        fb.get("theme"),
                "customizeTitle": fb.get("customizeTitle", ""),
                "faq":          faq_translated,
            }
            bg_id = strip_media_to_id(fb.get("backgroundImage"))
            if bg_id:
                entry["backgroundImage"] = {"id": bg_id}
            result.append(entry)

        elif comp == "feature.step-cards":
            # 从英文版重建（同 specific-features 模式），翻译 title/subtitle/steps
            en_sc = next((b for b in en_blocks if b["__component"] == "feature.step-cards"), None)
            steps_translated = strip_steps(en_sc.get("steps", []) if en_sc else [], t_map)
            result.append({
                "id":        bid,
                "__component": comp,
                "title":     translate((en_sc or {}).get("title"), t_map),
                "subtitle":  translate((en_sc or {}).get("subtitle"), t_map),
                "theme":     fb.get("theme"),
                "steps":     steps_translated,
            })

        else:
            # 其他 block：只传 id 和 __component，保持不变
            result.append({"id": bid, "__component": comp})

    return result


def verify(fr_page_id: int, en_blocks: list, locale: str, t_map: dict):
    """
    校验新建的本地化版本：
    - 拉取所有字段
    - 逐 block 对比英文原文，找出仍为英文的文本字段
    - 打印 PASS / WARN 报告
    """
    import re as _re

    resp = requests.get(
        f"{CMS_BASE}/api/special-topic-pages/{fr_page_id}"
        "?populate[blocks][populate][buttons][populate]=*"
        "&populate[blocks][populate][header][populate][icon]=*"
        "&populate[blocks][populate][trustedBy][populate]=*"
        "&populate[blocks][populate][faq]=*"
        "&populate[blocks][populate][subFeatures][populate]=*"
        "&populate[blocks][populate][steps][populate]=*"
        f"&locale={locale}",
        headers=headers(),
    )
    fr_attrs = resp.json()["data"]["attributes"]
    fr_blocks = fr_attrs.get("blocks", [])

    # 英文版文本字段集合（用于判断某字段是否"仍为英文"）
    en_texts: set = set()
    def collect_texts(obj):
        if isinstance(obj, dict):
            for k, v in obj.items():
                # label 是 UI 固定标签（如 "Trusted by"），不纳入检查
                if k in ("title", "text", "subtitle", "customizeText",
                         "question", "answer", "customizeTitle", "topTitle") and isinstance(v, str) and v.strip():
                    en_texts.add(_re.sub(r'\s+', ' ', v.strip()))
                collect_texts(v)
        elif isinstance(obj, list):
            for v in obj:
                collect_texts(v)
    for b in en_blocks:
        collect_texts(b)

    warnings = []

    def check_field(comp, field_path, value):
        if not value or not isinstance(value, str):
            return
        plain = _re.sub(r'<[^>]+>', '', value).strip()
        plain_norm = _re.sub(r'\s+', ' ', plain)
        if plain_norm in en_texts:
            warnings.append(f"  WARN [{comp}] {field_path} 仍为英文: {plain_norm[:60]}")

    for fb in fr_blocks:
        comp = fb.get("__component", "")

        # 通用文本字段
        for field in ("title", "label", "text", "subtitle"):
            check_field(comp, field, fb.get(field))

        # header
        if fb.get("header"):
            h = fb["header"]
            for field in ("title", "label", "text"):
                check_field(comp, f"header.{field}", h.get(field))

        # trustedBy（只检查 label）
        for i, tb in enumerate(fb.get("trustedBy", [])):
            if not tb.get("icon", {}).get("data"):
                warnings.append(f"  WARN [{comp}] trustedBy[{i}].icon 缺失 ({tb.get('label')})")

        # faq
        for i, item in enumerate(fb.get("faq", [])):
            check_field(comp, f"faq[{i}].question", item.get("question"))
            check_field(comp, f"faq[{i}].answer", item.get("answer"))

        # subFeatures
        for i, sf in enumerate(fb.get("subFeatures", [])):
            check_field(comp, f"subFeatures[{i}].title", sf.get("title"))
            check_field(comp, f"subFeatures[{i}].text", sf.get("text"))
            if not sf.get("media", {}).get("data"):
                warnings.append(f"  WARN [{comp}] subFeatures[{i}].media 缺失 ({sf.get('title','')})")

        # steps
        for i, s in enumerate(fb.get("steps", [])):
            check_field(comp, f"steps[{i}].customizeText", s.get("customizeText"))

        # cta buttons link（只检查 blocks.cta，feature-hero 的 button 结构不同）
        if comp == "blocks.cta":
            for i, btn in enumerate(fb.get("buttons", [])):
                if btn.get("link") is None:
                    warnings.append(f"  WARN [{comp}] buttons[{i}].link 为 null")

        # trust-by label 是固定 UI 标签，不检查

    if warnings:
        print(f"  发现 {len(warnings)} 个问题：")
        for w in warnings:
            print(w)
    else:
        print(f"  ✓ 全部通过，无未翻译字段")

    return len(warnings) == 0


def _find_locale_id(page_id: int, locale: str):
    """查询 page_id 下已存在的 locale 版本 id，返回 int 或 None。"""
    resp = requests.get(
        f"{CMS_BASE}/api/special-topic-pages/{page_id}",
        headers=headers(),
        params={"populate": "localizations"},
    )
    if resp.status_code != 200:
        return None
    locs = resp.json().get("data", {}).get("attributes", {}).get("localizations", {}).get("data", [])
    for loc in locs:
        if loc.get("attributes", {}).get("locale") == locale:
            return loc.get("id")
    return None


def localize(page_id: int, locale: str, sheet_name: str, publish: bool = True, force_truncate: bool = False):
    print(f"\n=== localize_special_topic_page ===")
    print(f"  page_id        : {page_id}")
    print(f"  locale         : {locale}")
    print(f"  sheet          : {sheet_name}")
    print(f"  publish        : {publish}")
    print(f"  force_truncate : {force_truncate}")
    print()

    # 1. 拉取英文版
    print("[1/5] 拉取英文版结构...")
    en_attrs = fetch_en_page(page_id)
    en_blocks = en_attrs.get("blocks", [])
    print(f"  blocks 数量: {len(en_blocks)}")

    # 2. 读取翻译表
    print("[2/5] 读取 Excel 翻译表...")
    t_map = build_translation_map(sheet_name, locale)

    # 3. 构建 POST payload
    print("[3/5] 构建 payload...")
    new_blocks = []
    for block in en_blocks:
        converted = convert_block(block, t_map, locale, force_truncate=force_truncate)
        new_blocks.append(converted)
        print(f"  ✓ {block['__component']}  →  {converted.get('title', '')[:50] or '(no title)'}")

    payload = {
        "locale":      locale,
        "slug":        en_attrs["slug"],  # 与英文版保持一致
        "navbarStyle": en_attrs.get("navbarStyle"),
        "blocks":      new_blocks,
    }

    # 4. POST 创建 localization（若已存在则 PUT 覆盖）
    print(f"\n[4/5] 创建 {locale} localization...")
    url = f"{CMS_BASE}/api/special-topic-pages/{page_id}/localizations"
    resp = requests.post(url, headers=headers_json(), json=payload)

    if resp.status_code not in (200, 201):
        err = resp.text
        if "locale is already used" in err:
            # 已存在该 locale，找到它的 id 然后 PUT 覆盖
            print(f"  locale 已存在，改为 PUT 覆盖...")
            existing_id = _find_locale_id(page_id, locale)
            if not existing_id:
                raise RuntimeError(f"locale {locale} 已存在但找不到其 id")
            put_resp = requests.put(
                f"{CMS_BASE}/api/special-topic-pages/{existing_id}",
                headers=headers_json(),
                json={"data": {**payload, "locale": locale}},
            )
            if put_resp.status_code != 200:
                raise RuntimeError(f"PUT {put_resp.status_code}: {put_resp.text[:300]}")
            new_id = existing_id
            print(f"  覆盖成功！id = {new_id}")
        else:
            print(f"  ERROR {resp.status_code}: {err}")
            raise RuntimeError(f"POST {resp.status_code}: {err[:300]}")
    else:
        result = resp.json()
        new_id = result.get("id")
        print(f"  创建成功！新 id = {new_id}, slug = {result.get('slug')}, locale = {result.get('locale')}")

    # 5. PUT 补全 trustedBy（POST 时无法传 trustedBy，需单独 PUT）
    print(f"[5/5] 补全 trustedBy 并发布...")
    fr_blocks = fetch_fr_blocks(new_id)
    patch_blocks = build_patch_blocks(fr_blocks, en_blocks, t_map, force_truncate=force_truncate)

    from datetime import datetime, timezone
    published_at = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.000Z") if publish else None
    patch_payload = {"data": {"blocks": patch_blocks}}
    if published_at:
        patch_payload["data"]["publishedAt"] = published_at

    patch_resp = requests.put(
        f"{CMS_BASE}/api/special-topic-pages/{new_id}",
        headers=headers_json(),
        json=patch_payload,
    )
    if patch_resp.status_code == 200:
        patch_data = patch_resp.json()["data"]["attributes"]
        print(f"  补全成功！publishedAt = {patch_data.get('publishedAt')}")
    else:
        print(f"  PATCH ERROR {patch_resp.status_code}: {patch_resp.text}")

    # 6. 校验：对比英文版逐字段检查是否仍为英文原文
    print(f"\n[6/5] 校验翻译结果...")
    verify(new_id, en_blocks, locale, t_map)

    print(f"\n完成。")
    print(f"  CMS 后台: {CMS_BASE}/admin/content-manager/collectionType/api::special-topic-page.special-topic-page/{new_id}?plugins[i18n][locale]={locale}")
    return new_id


def _localize_with_tmap(
    page_id: int,
    locale: str,
    en_attrs: dict,
    en_blocks: list,
    t_map: dict,
    publish: bool = True,
) -> int:
    """
    使用已构建好的 t_map 执行本地化（跳过 Excel 读取）。
    供 localize_agent.py 的 AI 翻译模式调用。
    返回新建页面的 id。
    """
    from datetime import datetime, timezone as _tz

    print(f"\n=== _localize_with_tmap ===")
    print(f"  page_id : {page_id}")
    print(f"  locale  : {locale}")
    print(f"  t_map   : {len(t_map)} 条")

    # 构建 POST payload（复用 localize() 内的逻辑）
    en_slug = en_attrs.get("slug", "")
    slug = en_slug  # 与英文版保持一致
    title_en = en_attrs.get("title") or ""

    blocks_payload = [convert_block(b, t_map, locale='en', force_truncate=True) for b in en_blocks]

    post_payload = {
        "locale": locale,
        "slug":   slug,
        "title":  translate(title_en, t_map) if title_en else slug,
        "blocks": blocks_payload,
    }

    resp = requests.post(
        f"{CMS_BASE}/api/special-topic-pages/{page_id}/localizations",
        headers=headers_json(),
        json=post_payload,
    )
    if resp.status_code not in (200, 201):
        raise RuntimeError(f"POST failed {resp.status_code}: {resp.text[:200]}")

    result = resp.json()
    new_id = result.get("id")
    print(f"  创建成功！新 id = {new_id}")

    # PUT 补全 trustedBy 并发布
    fr_blocks   = fetch_fr_blocks(new_id)
    patch_blocks = build_patch_blocks(fr_blocks, en_blocks, t_map, force_truncate=True)

    published_at = datetime.now(_tz.utc).strftime("%Y-%m-%dT%H:%M:%S.000Z") if publish else None
    patch_payload = {"data": {"blocks": patch_blocks}}
    if published_at:
        patch_payload["data"]["publishedAt"] = published_at

    patch_resp = requests.put(
        f"{CMS_BASE}/api/special-topic-pages/{new_id}",
        headers=headers_json(),
        json=patch_payload,
    )
    if patch_resp.status_code != 200:
        raise RuntimeError(f"PUT failed {patch_resp.status_code}: {patch_resp.text[:200]}")

    print(f"  补全成功！publishedAt = {patch_resp.json()['data']['attributes'].get('publishedAt')}")

    verify(new_id, en_blocks, locale, t_map)

    return new_id


# ── CLI 入口 ──────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="将 special-topic-page 英文版本地化为指定语言"
    )
    parser.add_argument("page_id",    type=int, help="英文版页面 ID（如 5）")
    parser.add_argument("locale",     type=str, help="目标语言代码（如 fr、de、es）")
    parser.add_argument("sheet_name", type=str, help="Excel sheet 名称（如 'Organize PDF'）")
    parser.add_argument("--no-publish", action="store_true", help="创建后不自动发布（保持 draft）")
    args = parser.parse_args()

    localize(
        page_id    = args.page_id,
        locale     = args.locale,
        sheet_name = args.sheet_name,
        publish    = not args.no_publish,
    )
